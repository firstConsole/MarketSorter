import pandas as pd
import openpyxl
import os
from datetime import datetime
from openpyxl.styles import Alignment, Font, Border, Side


def format_ozon_file(file):
    try:
        df = pd.read_csv(file, delimiter=';')
        print(f"Количество колонок после загрузки: {df.shape[1]}")

        columns_to_remove = [0, 2, 4, 5, 6, 7, 8, 9, 10, 12, 13, 14, 15, 17, 18, 19, 20, 21, 22, 23]
        df_filtered = df.drop(df.columns[columns_to_remove], axis=1, errors='ignore')

        df_filtered['Дата отгрузки'] = pd.to_datetime(df_filtered['Дата отгрузки'])
        df_filtered['Дата отгрузки'] = df_filtered['Дата отгрузки'].dt.strftime('%d.%m.%Y')

        print(f"DataFrame успешно извлечён: {df_filtered}")

        final_excel_file = filter_and_sort(df_filtered)

        return final_excel_file

    except Exception as e:
        print(f"Ошибка при обработке файла Ozon: {e}")
        return None


def filter_and_sort(df):
    now = datetime.now()
    timestamp = now.strftime('%d-%m-%Y_%H-%M-%S')

    header = df.columns.tolist()  # Сохранение заголовков
    data_with_header = df.reset_index(drop=True)  # Сброс индексов для корректной сортировки

    data_with_header['Дубликаты'] = data_with_header.duplicated(subset="Номер отправления", keep=False)

    sorted_data = data_with_header.sort_values(by="Артикул", ascending=True)

    result_df = pd.DataFrame(columns=header + ["Дубликаты"])
    result_df = pd.concat([result_df, sorted_data], ignore_index=True)

    print(result_df)
    result_file_path = f"OZON_{timestamp}_temp.xlsx"
    result_df.to_excel(result_file_path, index=False)

    wb = openpyxl.load_workbook(result_file_path)
    ws = wb.active

    ws.insert_rows(1)
    ws['A1'] = "OZON SELLER"

    title_font = Font(size=14, bold=True, italic=True)
    title_alignment = Alignment(horizontal='left', vertical='center')
    ws['A1'].font = title_font
    ws['A1'].alignment = title_alignment

    font = Font(size=14)
    alignment = Alignment(horizontal='center', vertical='center')

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        if row[0].row > 1:
            for cell in row:
                cell.font = font
                cell.alignment = alignment
                cell.border = thin_border


    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 15

    formatted_file_path = f"OZON_{timestamp}.xlsx"
    wb.save(formatted_file_path)

    if os.path.exists(result_file_path):
        os.remove(result_file_path)
        print(f"Временный файл {result_file_path} успешно удален.")
    else:
        print(f"Временный файл {result_file_path} не найден для удаления.")

    return formatted_file_path
