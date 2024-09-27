import openpyxl
import pandas as pd
import os
from openpyxl.styles import Alignment, Font, Border, Side
from datetime import datetime


def format_wb_file(file):
    try:
        now = datetime.now()
        timestamp = now.strftime('%d-%m-%Y_%H-%M-%S')

        df = pd.read_excel(file, header=None)
        df = df.iloc[2:].reset_index(drop=True)

        df.columns = df.iloc[0]
        df = df[1:].reset_index(drop=True)

        columns_to_remove = [1, 4, 5, 6]
        df_filtered = df.drop(df.columns[columns_to_remove], axis=1, errors='ignore')

        headers = ["№ задания", "Артикул продавца", "Стикер"]
        df_filtered.columns = headers

        article_sorted_path = os.path.join(os.getcwd(), f"WB_article_sort_{timestamp}.xlsx")
        stickers_sorted_path = os.path.join(os.getcwd(), f"WB_sticker_sort_{timestamp}.xlsx")

        make_article_sort(df_filtered, article_sorted_path)
        format_excel_file(article_sorted_path)

        if 'Стикер' in df_filtered.columns:
            df_filtered['Стикер'] = df_filtered['Стикер'].astype(str)
            df_filtered['Подписать'] = df_filtered['Стикер'].str[-4:]
        else:
            print("Ошибка: Колонка 'Стикер' не найдена.")
            return None

        make_stickers_sort(df_filtered, stickers_sorted_path)
        format_excel_file(stickers_sorted_path)

        return article_sorted_path, stickers_sorted_path

    except Exception as e:
        print(f"Произошла ошибка: {e}")
        return None


def make_stickers_sort(df, output_file):
    sorted_data = df.sort_values(by='Подписать', ascending=True)
    sorted_data.to_excel(output_file, index=False)

    return sorted_data


def make_article_sort(df, output_file):
    sorted_data = df.sort_values(by='Артикул продавца', ascending=True)
    sorted_data.to_excel(output_file, index=False)

    return sorted_data


def format_excel_file(file_path):
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        ws.insert_rows(1)
        ws['A1'] = "WILDBERRIES"

        title_font = Font(size=14, bold=True, italic=True)
        title_alignment = Alignment(horizontal='left', vertical='center')
        ws['A1'].font = title_font
        ws['A1'].alignment = title_alignment

        font = Font(size=14)
        alignment = Alignment(horizontal='center', vertical='center')

        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            if row[0].row > 1:
                for cell in row:
                    cell.font = font
                    cell.alignment = alignment
                    cell.border = thin_border


        ws.column_dimensions['A'].width = 33
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 18

        wb.save(file_path)
        print(f"Файл {file_path} успешно отформатирован.")

    except Exception as e:
        print(f"Ошибка при форматировании файла {file_path}: {e}")
