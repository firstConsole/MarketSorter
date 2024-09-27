import openpyxl
import pandas as pd
import os
from datetime import datetime
from openpyxl.styles import Alignment, Font, Border, Side


def format_yandex_file(file):
    now = datetime.now()
    timestamp = now.strftime('%d-%m-%Y_%H-%M-%S')
    df = pd.read_excel(file, header=None)
    df = df.iloc[1:].reset_index(drop=True)
    df_filtered = df[~df.apply(lambda row: row.astype(str).str.contains("Отменён в процессе обработки").any(), axis=1)]
    columns_to_remove = [1, 2, 4, 6, 7, 8, 9, 10, 13]
    df_filtered = df_filtered.drop(df_filtered.columns[columns_to_remove], axis=1, errors='ignore')
    duplicate_mask = df_filtered[0].duplicated(keep=False)
    df_filtered["Дубликаты"] = duplicate_mask

    output_file_path = f"Yandex_{timestamp}.xlsx"
    df_filtered.to_excel(output_file_path, index=False, header=False)

    sorted_file_path = sort_by_column(output_file_path, sort_column_index=1)
    final_excel_file = merge_repeated_cells(sorted_file_path)

    return final_excel_file


def merge_repeated_cells(file_path, column_index=0):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    ws.insert_rows(1)
    ws['A1'] = "ЯНДЕКС МАРКЕТ"

    title_font = Font(size=14, bold=True, italic=True)
    title_alignment = Alignment(horizontal='left', vertical='center')
    ws['A1'].font = title_font
    ws['A1'].alignment = title_alignment

    prev_value = None
    merge_start_row = None

    for row in range(3, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=column_index + 1).value

        if cell_value == prev_value:
            if merge_start_row is None:
                merge_start_row = row - 1
        else:
            if merge_start_row is not None:
                ws.merge_cells(start_row=merge_start_row, start_column=column_index + 1, end_row=row - 1,
                               end_column=column_index + 1)
                merge_start_row = None
            prev_value = cell_value

    if merge_start_row is not None:
        ws.merge_cells(start_row=merge_start_row, start_column=column_index + 1, end_row=ws.max_row,
                       end_column=column_index + 1)

    font = Font(size=14)
    alignment = Alignment(horizontal='center', vertical='center')

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        if row[0].row > 1:
            for cell in row:
                cell.font = font
                cell.alignment = alignment
                cell.border = thin_border

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 17
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 22
    ws.column_dimensions['F'].width = 10

    wb.save(file_path)
    return file_path


def sort_by_column(file_path, sort_column_index=1):
    df = pd.read_excel(file_path, header=None)

    header = df.iloc[0]
    data = df[1:]

    data_sorted = data.sort_values(by=[sort_column_index], ascending=True)

    df_sorted = pd.concat([header.to_frame().T, data_sorted], ignore_index=True)
    df_sorted.to_excel(file_path, index=False, header=False)

    print(f"File sorted by column {sort_column_index + 1} and saved to {file_path}")

    return file_path


def clean_up_temp_file(file_path):
    if os.path.exists(file_path):
        os.remove(file_path)
        print(f"Временный файл {file_path} успешно удалён.")
    else:
        print(f"Временный файл {file_path} не найден для удаления.")

